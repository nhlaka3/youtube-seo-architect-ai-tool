#!/usr/bin/env python3
"""Step 3: Build the multi-tab keyword XLSX from keyword data JSON.

Segments into the 4 master worksheets per the Marketing Brain spec:
  High Opportunity | Hidden Gems | High Volume | All Keywords

Accepts BOTH source formats:
  - ranked-keywords-<date>.json (DataForSEO: dict domain -> [kw rows])
  - keywords-mined-<date>.json  (free autocomplete mine: {keywords:[...]})
Priority: ranked-keywords (real volumes) > mined (demand proxy).

Input:  marketing/dataforseo/ranked-keywords-<date>.json  OR
        marketing/dataforseo/keywords-mined-<date>.json
Output: marketing/dataforseo/keyword-master-<date>.xlsx
"""
import json, sys
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "marketing" / "dataforseo"
today = date.today().isoformat()

# find latest ranked-keywords file, else latest mined file
files = sorted(DATA.glob("ranked-keywords-*.json"))
src_kind = "ranked"
if not files:
    files = sorted(DATA.glob("keywords-mined-*.json"))
    src_kind = "mined"
if not files:
    print("No ranked-keywords-*.json or keywords-mined-*.json found.")
    print("Run: node scripts/dataforseo-extract.mjs --keywords   OR")
    print("     python3 scripts/check-keyword.py --mine")
    sys.exit(1)
src = files[-1]
print(f"Reading {src.name} ({src_kind} source)")
raw = json.loads(src.read_text(encoding="utf-8"))

rows = []
if src_kind == "ranked":
    for domain, kws in raw.items():
        for k in kws:
            rows.append({
                "keyword": k.get("keyword", ""),
                "volume": k.get("volume", 0) or 0,
                "difficulty": k.get("difficulty", 0) or 0,
                "cpc": k.get("cpc", 0) or 0,
                "intent": k.get("intent", "") or "",
                "competitor": k.get("competitor", domain),
                "position": k.get("position", 0) or 0,
            })
else:  # mined: demand score is the volume proxy
    for k in raw.get("keywords", []):
        rows.append({
            "keyword": k.get("keyword", ""),
            "volume": k.get("demand", 0) or 0,          # demand proxy 0-100
            "difficulty": 0,                             # unknown in free mine
            "cpc": 0,
            "intent": "",
            "competitor": "/".join(k.get("sources", [])),
            "position": 0,
        })

# dedupe by keyword keeping the lowest difficulty / highest volume
seen = {}
for r in rows:
    kw = r["keyword"].lower()
    if kw not in seen or (r["volume"] > seen[kw]["volume"]):
        seen[kw] = r
rows = list(seen.values())
print(f"Unique keywords: {len(rows)}")

# ── Segmentation ────────────────────────────────────────────────
if src_kind == "ranked":
    high_opportunity = [r for r in rows if r["volume"] >= 100 and r["difficulty"] <= 30]
    hidden_gems      = [r for r in rows if 10 <= r["volume"] < 100 and r["difficulty"] <= 25]
    high_volume      = [r for r in rows if r["volume"] >= 1000]
else:  # mined: demand proxy 0-100
    high_opportunity = [r for r in rows if r["volume"] >= 70]      # high demand band
    hidden_gems      = [r for r in rows if 40 <= r["volume"] < 70] # mid demand
    high_volume      = []                                          # no real volume data
# All Keywords = everything not in the top three buckets
bucketed = {r["keyword"] for r in high_opportunity + hidden_gems + high_volume}
all_keywords     = [r for r in rows if r["keyword"] not in bucketed]

segments = [
    ("High Opportunity", high_opportunity, "High demand — immediate targets"),
    ("Hidden Gems", hidden_gems, "Mid-demand terms — instant rank potential"),
    ("High Volume", high_volume, "Top-of-funnel broad terms"),
    ("All Keywords", all_keywords, "Total corpus for long-tail authority"),
]

wb = Workbook()
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="0B6E99")
thin = Alignment(vertical="top")

for name, data, desc in segments:
    ws = wb.create_sheet(title=name[:31])
    ws.append(["Keyword", "Volume", "Difficulty", "CPC", "Intent", "Competitor", "Position"])
    for c in range(1, 8):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
    for r in sorted(data, key=lambda x: -x["volume"]):
        ws.append([r["keyword"], r["volume"], r["difficulty"], r["cpc"],
                   r["intent"], r["competitor"], r["position"]])
    for col, w in zip(range(1, 8), [55, 10, 10, 8, 12, 22, 8]):
        ws.column_dimensions[get_column_letter(col)].width = w
    print(f"  {name}: {len(data)} rows — {desc}")

# drop the default sheet, save
wb.remove(wb["Sheet"])
out = DATA / f"keyword-master-{today}.xlsx"
wb.save(out)
print(f"\n✅ {out.name} ({out.stat().st_size//1024} KB) — 4 tabs, {len(rows)} total keywords")
